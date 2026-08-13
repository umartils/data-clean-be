"""
Baca file upload (csv/xls/xlsx) jadi DataFrame, dan tulis DataFrame
hasil klasifikasi jadi satu file .xlsx dengan sheet terpisah per kategori
+ sheet gabungan "Hasil Klasifikasi" — mengikuti pola notebook asli.
"""

from __future__ import annotations

import io
import re
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

ID_LIKE_COLUMN_HINTS = ["id", "kode", "no.", "nomor", "npwp", "nik"]


def _is_id_like_column(col_name: str) -> bool:
    name_lower = str(col_name).strip().lower()
    return any(
        re.search(rf"(^|[^a-z0-9]){re.escape(hint)}([^a-z0-9]|$)", name_lower)
        for hint in ID_LIKE_COLUMN_HINTS
    )


def _peek_columns(file: BinaryIO, filename: str, csv_delimiter: str, csv_header_row: int) -> list[str]:
    """Baca header saja (0 baris data) untuk tahu nama kolom sebelum baca penuh."""
    lower = filename.lower()
    file.seek(0)
    if lower.endswith(".csv"):
        cols = pd.read_csv(file, delimiter=csv_delimiter, header=csv_header_row, nrows=0).columns
    else:
        engine = "openpyxl" if lower.endswith(".xlsx") else "xlrd"
        cols = pd.read_excel(file, sheet_name=0, engine=engine, nrows=0).columns
    file.seek(0)
    return list(cols)

def read_uploaded_file(
    file: BinaryIO,
    filename: str,
    sheet_name: str | int = 0,
    csv_delimiter: str = ";",
    csv_header_row: int = 0,
) -> pd.DataFrame:
    
    lower = filename.lower()
    columns = _peek_columns(file, filename, csv_delimiter, csv_header_row)
    dtype_map = {c: str for c in columns if _is_id_like_column(c)}

    if lower.endswith(".csv"):
        return pd.read_csv(file, delimiter=csv_delimiter, header=csv_header_row, dtype=dtype_map or None)
    if lower.endswith((".xls", ".xlsx")):
        engine = "openpyxl" if lower.endswith(".xlsx") else "xlrd"
        return pd.read_excel(file, sheet_name=sheet_name, engine=engine, dtype=dtype_map or None)
    raise ValueError(f"Format file tidak didukung: {filename}. Gunakan .csv, .xls, atau .xlsx")


def _sanitize_sheet_name(name: str, existing_names: list[str]) -> str:
    name = str(name)
    name = re.sub(r"[\\/?*\[\]:]", "-", name).strip() or "Tanpa Kategori"
    name = name[:31]
    original, counter = name, 2
    while name.lower() in {n.lower() for n in existing_names}:
        suffix = f" ({counter})"
        name = original[: 31 - len(suffix)] + suffix
        counter += 1
    return name


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4


def build_classified_workbook(
    df_original: pd.DataFrame,
    df_classified: pd.DataFrame,
    sheet_name_original: str = "Data Transaksi",
) -> bytes:
    """
    Bangun satu file .xlsx berisi:
      - sheet data asli
      - satu sheet per kategori (kolom 'Kategori')
      - sheet gabungan 'Hasil Klasifikasi' (semua baris + kolom Kategori & Kelompok Nominal)

    Return bytes, siap dikirim sebagai StreamingResponse.
    """
    buffer = io.BytesIO()
    existing_names = [sheet_name_original]
    kategori_ke_sheet: dict[str, str] = {}

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_original.to_excel(writer, sheet_name=sheet_name_original, index=False)

        for kategori in df_classified["Kategori"].unique():
            subset = df_classified[df_classified["Kategori"] == kategori]
            sheet_name = _sanitize_sheet_name(kategori, existing_names)
            existing_names.append(sheet_name)
            kategori_ke_sheet[kategori] = sheet_name
            subset.to_excel(writer, sheet_name=sheet_name, index=False)

        df_classified.to_excel(writer, sheet_name="Hasil Klasifikasi", index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    for sheet_name in [*kategori_ke_sheet.values(), "Hasil Klasifikasi"]:
        _style_header(wb[sheet_name])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
